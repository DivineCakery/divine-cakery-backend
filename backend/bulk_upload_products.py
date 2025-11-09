#!/usr/bin/env python3
"""
Bulk Product Upload Script for Divine Cakery
Uploads products from Excel file with optional images
"""

import openpyxl
import base64
import os
import sys
from pathlib import Path
import asyncio
import httpx
from dotenv import load_dotenv

load_dotenv()

# Configuration
EXCEL_FILE = "bulk_upload_items.xlsx"
IMAGES_FOLDER = "PHOTOS FOR PRODUCT LIST AND APP"
BACKEND_URL = "http://localhost:8001/api"

# Admin credentials (from create_demo_admins.py)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# Statistics
stats = {
    "total": 0,
    "success": 0,
    "failed": 0,
    "no_image": 0,
    "with_image": 0
}

failed_products = []


def get_image_base64(image_name):
    """Convert image to base64, returns None if not found"""
    if not image_name:
        return None
    
    # Try different extensions
    extensions = ['.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG']
    
    for ext in extensions:
        # Try exact match
        image_path = Path(IMAGES_FOLDER) / f"{image_name}{ext}"
        if image_path.exists():
            try:
                with open(image_path, 'rb') as f:
                    image_data = f.read()
                    base64_str = base64.b64encode(image_data).decode('utf-8')
                    return f"data:image/png;base64,{base64_str}"
            except Exception as e:
                print(f"Error reading image {image_path}: {e}")
                return None
        
        # Try case-insensitive match
        parent_dir = Path(IMAGES_FOLDER)
        if parent_dir.exists():
            for file in parent_dir.iterdir():
                if file.is_file() and file.stem.lower() == image_name.lower():
                    try:
                        with open(file, 'rb') as f:
                            image_data = f.read()
                            base64_str = base64.b64encode(image_data).decode('utf-8')
                            return f"data:image/png;base64,{base64_str}"
                    except Exception as e:
                        print(f"Error reading image {file}: {e}")
                        return None
    
    return None


async def login_admin(client):
    """Login and get admin token"""
    print("Logging in as admin...")
    response = await client.post(
        f"{BACKEND_URL}/login",
        data={"username": ADMIN_USERNAME, "password": ADMIN_PASSWORD}
    )
    
    if response.status_code == 200:
        data = response.json()
        token = data.get("access_token")
        print(f"✓ Login successful, token: {token[:20]}...")
        return token
    else:
        print(f"✗ Login failed: {response.status_code} - {response.text}")
        return None


async def create_product(client, token, product_data):
    """Create a single product"""
    headers = {"Authorization": f"Bearer {token}"}
    
    response = await client.post(
        f"{BACKEND_URL}/products",
        headers=headers,
        json=product_data,
        timeout=30.0
    )
    
    return response


def map_food_type(fssai_value):
    """Map FSSAI food type to veg/non-veg"""
    if not fssai_value:
        return "veg"
    
    fssai_str = str(fssai_value).lower().strip()
    if "non" in fssai_str or "non-veg" in fssai_str:
        return "non-veg"
    return "veg"


def clean_string(value):
    """Clean and return string value"""
    if value is None:
        return None
    return str(value).strip() if str(value).strip() else None


async def bulk_upload():
    """Main bulk upload function"""
    print("=" * 60)
    print("DIVINE CAKERY - BULK PRODUCT UPLOAD")
    print("=" * 60)
    
    # Load Excel file
    print(f"\nLoading Excel file: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    total_rows = sheet.max_row - 1  # Exclude header
    stats["total"] = total_rows
    print(f"Found {total_rows} products to upload\n")
    
    # Create HTTP client
    async with httpx.AsyncClient() as client:
        # Login
        token = await login_admin(client)
        if not token:
            print("Failed to login. Exiting.")
            return
        
        print(f"\nStarting upload of {total_rows} products...")
        print("-" * 60)
        
        # Process each row
        for row_idx in range(2, sheet.max_row + 1):
            row_num = row_idx - 1
            
            # Extract data from Excel
            product_image_name = clean_string(sheet.cell(row_idx, 1).value)
            product_name = clean_string(sheet.cell(row_idx, 2).value)
            category1 = clean_string(sheet.cell(row_idx, 3).value)
            category2 = clean_string(sheet.cell(row_idx, 4).value)
            mrp = sheet.cell(row_idx, 5).value
            selling_price = sheet.cell(row_idx, 6).value
            unit = clean_string(sheet.cell(row_idx, 7).value)
            shelf_life = clean_string(sheet.cell(row_idx, 8).value)
            storage_instructions = clean_string(sheet.cell(row_idx, 9).value)
            ingredients = clean_string(sheet.cell(row_idx, 10).value)
            allergen_info = clean_string(sheet.cell(row_idx, 11).value)
            description = clean_string(sheet.cell(row_idx, 12).value)
            packet_size = clean_string(sheet.cell(row_idx, 13).value)
            food_type_fssai = sheet.cell(row_idx, 14).value
            
            # Validate required fields
            if not product_name:
                print(f"[{row_num}/{total_rows}] ✗ Skipping row {row_idx}: No product name")
                stats["failed"] += 1
                failed_products.append({"row": row_idx, "name": "Unknown", "reason": "No product name"})
                continue
            
            # Build categories list
            categories = []
            if category1:
                categories.append(category1)
            if category2:
                categories.append(category2)
            
            # Use first category as primary category (for backward compatibility)
            primary_category = categories[0] if categories else "Uncategorized"
            
            # Get image
            image_base64 = None
            if product_image_name:
                image_base64 = get_image_base64(product_image_name)
                if image_base64:
                    stats["with_image"] += 1
                else:
                    stats["no_image"] += 1
            else:
                stats["no_image"] += 1
            
            # Map food type
            food_type = map_food_type(food_type_fssai)
            
            # Prepare product data
            product_data = {
                "name": product_name,
                "description": description,
                "category": primary_category,
                "categories": categories,
                "mrp": float(mrp) if mrp else 0.0,
                "price": float(selling_price) if selling_price else 0.0,
                "packet_size": packet_size,
                "unit": unit if unit else "piece",
                "remarks": None,
                "image_base64": image_base64,
                "is_available": True,
                "closing_stock": 0,
                "shelf_life": shelf_life,
                "storage_instructions": storage_instructions,
                "food_type": food_type,
                "ingredients": ingredients,
                "allergen_info": allergen_info
            }
            
            # Create product
            try:
                response = await create_product(client, token, product_data)
                
                if response.status_code == 200:
                    stats["success"] += 1
                    img_status = "📷" if image_base64 else "📄"
                    print(f"[{row_num}/{total_rows}] ✓ {img_status} {product_name[:40]}")
                else:
                    stats["failed"] += 1
                    error_msg = response.text[:100] if response.text else "Unknown error"
                    print(f"[{row_num}/{total_rows}] ✗ {product_name[:40]} - {response.status_code}: {error_msg}")
                    failed_products.append({
                        "row": row_idx,
                        "name": product_name,
                        "reason": f"{response.status_code}: {error_msg}"
                    })
                    
            except Exception as e:
                stats["failed"] += 1
                error_msg = str(e)[:100]
                print(f"[{row_num}/{total_rows}] ✗ {product_name[:40]} - Error: {error_msg}")
                failed_products.append({
                    "row": row_idx,
                    "name": product_name,
                    "reason": error_msg
                })
            
            # Small delay to avoid overwhelming the server
            await asyncio.sleep(0.1)
    
    # Print summary
    print("\n" + "=" * 60)
    print("UPLOAD SUMMARY")
    print("=" * 60)
    print(f"Total products: {stats['total']}")
    print(f"✓ Successfully uploaded: {stats['success']}")
    print(f"✗ Failed: {stats['failed']}")
    print(f"📷 With images: {stats['with_image']}")
    print(f"📄 Without images: {stats['no_image']}")
    
    if failed_products:
        print(f"\n⚠️  Failed Products ({len(failed_products)}):")
        print("-" * 60)
        for failed in failed_products[:10]:  # Show first 10
            print(f"Row {failed['row']}: {failed['name'][:40]} - {failed['reason']}")
        if len(failed_products) > 10:
            print(f"... and {len(failed_products) - 10} more")
    
    print("\n" + "=" * 60)
    
    return stats["success"] == stats["total"]


if __name__ == "__main__":
    # Run the upload
    success = asyncio.run(bulk_upload())
    sys.exit(0 if success else 1)
